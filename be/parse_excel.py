import openpyxl
import json
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\LENOVO\Downloads\Tabel implementasi RPAM sd Triwulan IV baru.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

# 1. Establish locations mappings
# Prepopulate with names from the seed
locations_map = {
    'C1': {'kodeLokasi': 'C1', 'simbol': '〇', 'namaLokasi': 'Klorinasi di Reservoir Kalipuro'},
    'A1': {'kodeLokasi': 'A1', 'simbol': '〇', 'namaLokasi': 'Mata Air Gedor II'},
    'A2': {'kodeLokasi': 'A2', 'simbol': '〇', 'namaLokasi': 'Sumur Pompa Ketapang'},
    'A3': {'kodeLokasi': 'A3', 'simbol': '〇', 'namaLokasi': 'Sumber Gedor I'},
    'R1': {'kodeLokasi': 'R1', 'simbol': '▽', 'namaLokasi': 'Reservoir Kalipuro'},
    'R2': {'kodeLokasi': 'R2', 'simbol': '▽', 'namaLokasi': 'Reservoir Banjarsari'},
    'D1': {'kodeLokasi': 'D1', 'simbol': '→', 'namaLokasi': 'Distribusi Kalongan'},
    'UP1': {'kodeLokasi': 'UP1', 'simbol': '→', 'namaLokasi': 'Unit Pelayanan Kel. Kalipuro'}
}

# Add fallback generation for other codes
def get_or_create_location(code):
    code = code.strip()
    if code in locations_map:
        return locations_map[code]
    
    # Generate based on prefix
    simbol = '〇'
    prefix = 'Air Baku'
    if code.startswith('A'):
        prefix = 'Air Baku'
        simbol = '〇'
    elif code.startswith('R'):
        prefix = 'Reservoir'
        simbol = '▽'
    elif code.startswith('D'):
        prefix = 'Distribusi'
        simbol = '→'
    elif code.startswith('UP'):
        prefix = 'Unit Pelayanan'
        simbol = '→'
    elif code.startswith('C'):
        prefix = 'Klorinasi'
        simbol = '〇'
        
    loc = {
        'kodeLokasi': code,
        'simbol': simbol,
        'namaLokasi': f"{prefix} {code}"
    }
    locations_map[code] = loc
    return loc

# Global database dict keyed by (kodeLokasi, kodeRisiko)
db_records = {}

# Helper to safely parse int
def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(val))
    except:
        return default

# Helper to clean string
def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

# --- 1. M3.5 Penilaian Resiko (contains both Identifikasi Bahaya and Penilaian Risiko) ---
sheet = wb['M3.5 Penilaian Resiko']
rows = list(sheet.iter_rows(values_only=True))

for idx in range(8, len(rows)):
    row = rows[idx]
    if len(row) > 3 and row[1] is not None and row[3] is not None:
        loc_code = clean_str(row[1])
        risk_code = clean_str(row[3])
        
        # Check if it's a valid row
        if not re.match(r'^(A|C|R|D|UP)\d+$', loc_code):
            continue
            
        get_or_create_location(loc_code)
        
        komponen_spam = clean_str(row[5])
        kontaminasi_x = clean_str(row[7])
        komponen_spam_y = clean_str(row[10])
        penyebab_z = clean_str(row[13])
        kejadian_xyz = clean_str(row[16])
        tipe_bahaya = clean_str(row[19])
        
        # Clean type bahaya
        tipe_bahaya_clean = 'Fisik'
        if 'fis' in tipe_bahaya.lower():
            tipe_bahaya_clean = 'Fisik'
        elif 'kim' in tipe_bahaya.lower():
            tipe_bahaya_clean = 'Kimia'
        elif 'mik' in tipe_bahaya.lower():
            tipe_bahaya_clean = 'Mikrobiologi'
        else:
            tipe_bahaya_clean = tipe_bahaya
            
        peluang = safe_int(row[22])
        dampak = safe_int(row[25])
        skor = safe_int(row[28])
        if skor == 0:
            skor = peluang * dampak
            
        key = (loc_code, risk_code)
        db_records[key] = {
            'kodeLokasi': loc_code,
            'kodeRisiko': risk_code,
            'identifikasi': {
                'komponenSpam': komponen_spam,
                'kontaminasiX': kontaminasi_x,
                'komponenSpamY': komponen_spam_y,
                'penyebabZ': penyebab_z,
                'kejadianBahayaXYZ': kejadian_xyz,
                'tipeBahaya': tipe_bahaya_clean
            },
            'penilaian': {
                'peluangKejadianBahaya': peluang,
                'dampakKeparahan': dampak,
                'skorRisiko': skor
            },
            'kajiUlang': None,
            'rencanaPerbaikan': None,
            'pemantauanOperasional': None
        }

print(f"Loaded {len(db_records)} records from M3.5 Penilaian Resiko")

# --- 2. M4 Kaji Ulang Risiko ---
sheet = wb['M4 Kaji Ulang Risiko']
rows = list(sheet.iter_rows(values_only=True))

for idx in range(7, len(rows)):
    row = rows[idx]
    if len(row) > 2 and row[1] is not None and row[2] is not None:
        loc_code = clean_str(row[1])
        risk_code = clean_str(row[2])
        key = (loc_code, risk_code)
        
        if key not in db_records:
            # Skip if we didn't find the hazard in M3.5
            continue
            
        tindakan = clean_str(row[13])
        referensi = clean_str(row[14])
        
        # Validasi
        val_efektif = row[15]
        val_tidak_efektif = row[16]
        val_tidak_pasti = row[17]
        
        validasi_enum = 'TIDAK_PASTI'
        if val_efektif is not None and clean_str(val_efektif).lower() in ['v', '1']:
            validasi_enum = 'EFEKTIF'
        elif val_tidak_efektif is not None and clean_str(val_tidak_efektif).lower() in ['v', '1']:
            validasi_enum = 'TIDAK_EFEKTIF'
        elif val_tidak_pasti is not None and clean_str(val_tidak_pasti).lower() in ['v', '1']:
            validasi_enum = 'TIDAK_PASTI'
            
        peluang = safe_int(row[18])
        dampak = safe_int(row[19])
        skor = safe_int(row[20])
        if skor == 0:
            skor = peluang * dampak
            
        # Only populate kajiUlang if there is a tindakanPengendalian or a validation
        if tindakan or referensi or val_efektif or val_tidak_efektif or val_tidak_pasti:
            db_records[key]['kajiUlang'] = {
                'tindakanPengendalian': tindakan if tindakan else "N/A",
                'referensi': referensi if referensi else "N/A",
                'validasi': validasi_enum,
                'peluangKejadianBahaya': peluang,
                'dampakKeparahan': dampak,
                'skorRisiko': skor
            }

kaji_count = sum(1 for r in db_records.values() if r['kajiUlang'] is not None)
print(f"Loaded {kaji_count} kaji ulang records from M4")

# --- 3. M5 Rencana Perbaikan ---
sheet = wb['M5 Rencana Perbaikan']
rows = list(sheet.iter_rows(values_only=True))

for idx in range(7, len(rows)):
    row = rows[idx]
    if len(row) > 2 and row[1] is not None and row[2] is not None:
        loc_code = clean_str(row[1])
        risk_code = clean_str(row[2])
        key = (loc_code, risk_code)
        
        if key not in db_records:
            continue
            
        rencana = clean_str(row[15])
        if not rencana:
            continue
            
        pj = clean_str(row[16])
        jadwal = clean_str(row[17])
        
        # Biaya
        biaya_raw = row[18]
        biaya = 0.0
        if biaya_raw is not None:
            try:
                biaya = float(biaya_raw)
            except:
                biaya = 0.0
                
        sumber = clean_str(row[19])
        
        # Normalisasi statusKemajuan ke enum
        status_raw = clean_str(row[20]).lower()
        if 'selesai' in status_raw or 'done' in status_raw:
            status_kemajuan = 'SELESAI'
        elif 'berjalan' in status_raw or 'proses' in status_raw or 'progress' in status_raw:
            status_kemajuan = 'SEDANG_BERJALAN'
        elif 'tunda' in status_raw or 'delay' in status_raw:
            status_kemajuan = 'TERTUNDA'
        else:
            status_kemajuan = 'BELUM_MULAI'

        # Gabungkan kendala menjadi satu field string
        kendala_parts = []
        if clean_str(row[21]).lower() in ['v', '1', 'ya', 'true']:
            kendala_parts.append('Keuangan')
        if clean_str(row[22]).lower() in ['v', '1', 'ya', 'true']:
            kendala_parts.append('Tenaga Kerja')
        kendala = ', '.join(kendala_parts) if kendala_parts else None
        
        prioritas = 'MENENGAH'
        if clean_str(row[23]).lower() in ['v', '1']:
            prioritas = 'PENDEK'
        elif clean_str(row[24]).lower() in ['v', '1']:
            prioritas = 'MENENGAH'
        elif clean_str(row[25]).lower() in ['v', '1']:
            prioritas = 'PANJANG'
            
        db_records[key]['rencanaPerbaikan'] = {
            'rencanaPerbaikan': rencana,
            'penanggungJawab': pj if pj else "N/A",
            'jadwal': jadwal if jadwal else "N/A",
            'biaya': biaya,
            'sumberPembiayaan': sumber if sumber else "N/A",
            'statusKemajuan': status_kemajuan,
            'kendala': kendala,
            'prioritas': prioritas
        }

rencana_count = sum(1 for r in db_records.values() if r['rencanaPerbaikan'] is not None)
print(f"Loaded {rencana_count} rencana perbaikan records from M5")

# --- 4. M6.2 Pemantauan Operasional ---
sheet = wb['M6.2 Pemantauan Operasional']
rows = list(sheet.iter_rows(values_only=True))

for idx in range(5, len(rows)):
    row = rows[idx]
    if len(row) > 2 and row[1] is not None and row[2] is not None:
        loc_code = clean_str(row[1])
        risk_code = clean_str(row[2])
        key = (loc_code, risk_code)
        
        if key not in db_records:
            continue
            
        monitor = clean_str(row[14])
        if not monitor:
            continue
            
        batas = clean_str(row[13])
        dimana = clean_str(row[15])
        kapan = clean_str(row[16])
        bagaimana = clean_str(row[17])
        pelaksana = clean_str(row[18])
        analis = clean_str(row[19])
        penerima = clean_str(row[20])
        koreksi = clean_str(row[21])
        pelaksana_kor = clean_str(row[22])
        waktu_kor = clean_str(row[23])
        penerima_kor = clean_str(row[24])
        
        db_records[key]['pemantauanOperasional'] = {
            'batasKritis': batas if batas else None,
            'apaYangDimonitor': monitor,
            'dimana': dimana if dimana else "N/A",
            'kapan': kapan if kapan else "N/A",
            'bagaimana': bagaimana if bagaimana else "N/A",
            'siapaYangMelakukan': pelaksana if pelaksana else "N/A",
            'siapaYangAkanMenganalisisHasilnya': analis if analis else "N/A",
            'siapaYangMenerimaHasilAnalisisDanMengambilTindakan': penerima if penerima else "N/A",
            'apaTindakanKoreksinya': koreksi if koreksi else "N/A",
            'siapaYangMelakukanTindakanKoreksi': pelaksana_kor if pelaksana_kor else "N/A",
            'seberapaCepat': waktu_kor if waktu_kor else "N/A",
            'siapaYangWajibMenerimaLaporanUntukTindakanKoreksiIni': penerima_kor if penerima_kor else "N/A"
        }

mon_count = sum(1 for r in db_records.values() if r['pemantauanOperasional'] is not None)
print(f"Loaded {mon_count} pemantauan operasional records from M6.2")

# Export to JSON
output_data = {
    'locations': list(locations_map.values()),
    'records': [v for v in db_records.values()]
}

with open('rpam_data.json', 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print("\nSaved rpam_data.json successfully!")
